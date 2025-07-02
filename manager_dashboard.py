import sqlite3
from flask import Flask, render_template, request, redirect, jsonify, url_for, flash, render_template_string
from flask_cors import CORS
import pandas as pd
from datetime import datetime
from jira import JIRA
import json
import urllib3
urllib3.disable_warnings()

token = "ODc4NDQzNTI2MTI4OghLCU8T4lrpVrkuqcfYMlM5vd6F"
jira_server = "https://jira.gtie.dell.com"
jira = JIRA(options={'server': jira_server, 'verify': False}, token_auth=token)

mapping_df = pd.read_excel("CET-CORE.xlsx")
mapping_df['PLATFORM'] = mapping_df['PLATFORM'].astype(str).str.strip().str.upper()
platform_to_corecet = dict(zip(mapping_df['PLATFORM'], mapping_df['CORE-CET']))

DB_PATH = "dashboard.db"
tracked_names = ["P_Hota", "Mohanraj_Chinnasamy", "Sakthi_Kumar", 
                 "Santhoshkumar_Kannap", "Aparajitha_Rajapur", "Nithyasri_Arava", "Darshan_M2", "Prakyat_Shetty", "Prajna_Harish"]

ENGINEERS = [
    ("P_Hota", "PHota@example.com"),
    ("Nithyasri_Arava", "Nithyasri_Arava@example.com"),

    ("Santoshkumar_Kannap", "Santhoshkumar_Kannap@example.com"),
    ("Darshan_M2", "Darshan_M2@example.com"),
    ("Prakyat_Shetty", "Prakyat_Shetty@example.com"),
    ("Prajna_Harish", "Prajna_Harish@example.com"),
    ("Karrthik_C_R", "Karrthik_C@example.com"),
    ("Ram_Yerra","ram_yerra@example.com"),
    ("Ganeshkumar Karthikeyan","Ganesh_Karthikeyan@example.com")
]

MANAGERS = [
    ("Mohanraj_Chinnasamy", "Mohanraj_Chinnasamy@example.com", 1),
    ("Pravin_Janakiram", "pravin@example.com", 2),
]

TECHNICIANS = [
    ("Dayalkumar_S", "Dayalkumar@example.com"),
    ("Shreyank_Shanbhag", "shreyank@example.com"),
    
]
#DATE has been replaced by DATETIME
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS managers (
      manager_id   INTEGER PRIMARY KEY,
      name         TEXT  UNIQUE NOT NULL,
      email        TEXT
    );
    CREATE TABLE IF NOT EXISTS engineers (
      engineer_id  INTEGER PRIMARY KEY AUTOINCREMENT,
      name         TEXT  UNIQUE NOT NULL,
      email        TEXT
    );
    CREATE TABLE IF NOT EXISTS technician (
      technician_id  INTEGER PRIMARY KEY AUTOINCREMENT,
      name         TEXT  UNIQUE NOT NULL,
      email        TEXT
    );
    CREATE TABLE IF NOT EXISTS issues (
      issue_id      INTEGER PRIMARY KEY AUTOINCREMENT,
      jit_number    TEXT    UNIQUE NOT NULL,
      summary       TEXT,
      status        TEXT,
      created_date  DATETIME DEFAULT CURRENT_TIMESTAMP,
      closed_date   DATETIME,
      close_requested_by TEXT,
      close_requested_date DATETIME,
      manager_id    INTEGER,
      engineer_id   INTEGER,
      technician_id INTEGER,
      assigned_date DATETIME,
      delete_issue  INTEGER DEFAULT 0,
      severity INTEGER,
      core_cet TEXT,
      platform TEXT,
      rce_assign_date DATETIME,
      generation TEXT,
      damage_found TEXT,
      damage_introduced TEXT,
      dam_used TEXT,
      new_jit_add_date DATETIME
      engineer_history TEXT,
      reject_request INTEGER DEFAULT 0,
      read1 INTEGER DEFAULT 0,
      read2 INTEGER DEFAULT 0,
      
      FOREIGN KEY(manager_id)   REFERENCES managers(manager_id),
      FOREIGN KEY(engineer_id)  REFERENCES engineers(engineer_id),
      FOREIGN KEY(technician_id) REFERENCES technician(technician_id)
    );
    CREATE TABLE IF NOT EXISTS notifications (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      engineer_id INTEGER,
      technician_id INTEGER,
      message TEXT,
      created_date DATETIME,
      read1 INTEGER DEFAULT 0,
      read2 INTEGER DEFAULT 0, 
      assigned_or_not INTEGER DEFAULT 0,
      FOREIGN KEY(engineer_id) REFERENCES engineers(engineer_id),
      FOREIGN KEY(technician_id) REFERENCES technician(technician_id)
    );
                     -- NEW: table of all possible damage types
    CREATE TABLE IF NOT EXISTS damage_types (
      id        INTEGER PRIMARY KEY AUTOINCREMENT,
      category  TEXT    NOT NULL,
      value     TEXT    NOT NULL
    );

    -- NEW: engineer picks for each issue
    CREATE TABLE IF NOT EXISTS engineer_damage (
      id          INTEGER PRIMARY KEY AUTOINCREMENT,
      issue_id    INTEGER NOT NULL,
      damage_id   INTEGER NOT NULL,
      introduced  INTEGER NOT NULL CHECK(introduced IN (0,1)),
      FOREIGN KEY(issue_id)  REFERENCES issues(issue_id),
      FOREIGN KEY(damage_id) REFERENCES damage_types(id)
    );

    -- NEW: manager comments
    CREATE TABLE IF NOT EXISTS manager_comments (
      id           INTEGER PRIMARY KEY AUTOINCREMENT,
      issue_id     INTEGER NOT NULL,
      comment      TEXT    NOT NULL,
      commented_at DATETIME DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(issue_id) REFERENCES issues(issue_id)
    );
  
                    
    """)
    # Add engineer_history column if it doesn't exist
    c.execute("PRAGMA table_info(issues)")
    columns = [col[1] for col in c.fetchall()]
    if 'engineer_history' not in columns:
        c.execute("ALTER TABLE issues ADD COLUMN engineer_history TEXT")

    for name, email, manager_id in MANAGERS:
        c.execute(
            "INSERT OR IGNORE INTO managers(name, email, manager_id) VALUES(?, ?, ?)",
            (name, email, manager_id)
        )
    for name, email in ENGINEERS:
        c.execute(
            "INSERT OR IGNORE INTO engineers(name, email) VALUES(?, ?)",
            (name, email)
        )
    for name, email in TECHNICIANS:
        c.execute(
            "INSERT OR IGNORE INTO technician(name, email) VALUES(?, ?)",
            (name, email)
        )

    damage_entries = [
        ('CPU', 'Socket pin bends'),
        ('CPU', 'Damages to CPU'),
        ('Slot', 'HPM'),
        ('Slot', 'any PCB slot damages'),
        ('Cable', 'Spraying'),
        ('Cable', 'continuity issues'),
        ('Cable', 'damages'),
        ('DIMM', 'Memory Module issue'),
        ('PCIe card', 'SNIC'),
        ('PCIe card', 'NIC'),
        ('PCIe card', 'DPU'),
        ('BP', 'Damage'),
        ('PERC', 'Damage'),
        ('GPU', 'Damage'),
        ('Power', 'PSU'),
        ('Power', 'Brick'),
    ]
    for cat, val in damage_entries:
        c.execute(
            "INSERT OR IGNORE INTO damage_types(category, value) VALUES(?, ?)",
            (cat, val)
        )

    conn.commit()
    conn.close()

def process_new_jits(new_jits_df: pd.DataFrame):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for _, row in new_jits_df.iterrows():
        jit = str(row["JIT Number"]).strip()
        summ = row["Summary"]
        stat = "Open"
        created = row["Created Date"]
        severity = row["Severity"]
        core_cet = row["CORE-CET"]
        platform = row["Platform"]
        rce_assign_date = row["RCE Assigned Date"]
        generation = row["Generation"]
        c.execute("""
            INSERT INTO issues(jit_number, summary, status, created_date, severity, core_cet, platform, rce_assign_date, generation, engineer_history)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '[]')
            ON CONFLICT(jit_number) DO UPDATE SET
                summary      = excluded.summary,
                status       = excluded.status,
                created_date = excluded.created_date,
                severity     = excluded.severity,
                core_cet     = excluded.core_cet,
                platform     = excluded.platform,
                rce_assign_date = excluded.rce_assign_date,
                generation   = excluded.generation
        """, (jit, summ, stat, created, severity, core_cet, platform, rce_assign_date, generation))
    
    c.execute("SELECT platform FROM issues")
    rows = c.fetchall()
    core_cet_df = pd.read_excel("CET-CORE.xlsx")
    core_cet_df['PLATFORM'] = core_cet_df['PLATFORM'].str.strip().str.upper()
    
    for row in rows:
        platform = row[0]
        matches = core_cet_df[core_cet_df['PLATFORM'] == platform]
        if platform == '-':
            matches["CORE-CET"] = '-'
        if not matches.empty:
            core_cet_values = matches['CORE-CET'].values
            for core_cet_value in core_cet_values:
                c.execute("UPDATE issues SET core_cet = ? WHERE platform = ?", (core_cet_value, platform))
            conn.commit()
    conn.close()

def get_engineer_loads():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
      SELECT e.engineer_id, COUNT(i.issue_id) AS open_count
      FROM engineers e
      LEFT JOIN issues i
        ON i.engineer_id = e.engineer_id
        AND i.status <> 'Closed'
      GROUP BY e.engineer_id
      ORDER BY open_count ASC
    """)
    eng_ids = [row[0] for row in c.fetchall()]
    conn.close()
    return eng_ids

def assign_unassigned_issues():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
      SELECT issue_id
      FROM issues
      WHERE status <> 'Closed' AND engineer_id IS NULL
    """)
    unassigned = [r[0] for r in c.fetchall()]
    eng_ids = get_engineer_loads()
    if not eng_ids:
        conn.close()
        return
    for idx, issue_id in enumerate(unassigned):
        eng = eng_ids[idx % len(eng_ids)]
        c.execute("""
          UPDATE issues
          SET engineer_id = ?, assigned_date = DATE('now')
          WHERE issue_id = ?
        """, (eng, issue_id))
    conn.commit()
    conn.close()

import sqlite3
from datetime import datetime, timedelta

# List of government holidays to skip (strings in 'YYYY-MM-DD' format)
HOLIDAYS = {
    
   '2025-01-14',
        '2025-2-25',
        '2025-3-31',
        '2025-5-1',
       '2025-8-15',
        '2025-8-27',
        '2025-10-1',
        '2025-10-2',
        '2025-10-20',
        '2025-12-25',
    # … add more …
}

def business_days_between(start_date: datetime, end_date: datetime, holidays: set) -> int:
    """
    Count days from start_date to end_date inclusive,
    excluding Saturdays (weekday=5), Sundays (weekday=6), and any date in holidays.
    """
    days = 0
    cur = start_date
    while cur <= end_date:
        if cur.weekday() < 5 and cur.strftime('%Y-%m-%d') not in holidays:
            days += 1
        cur += timedelta(days=1)
    return days

def compute_metrics():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 1. Fetch all CLOSED issues with their assigned and closed dates
    c.execute("""
        SELECT engineer_id, assigned_date, closed_date
        FROM issues
        WHERE status = 'Closed'
          AND assigned_date IS NOT NULL
          AND closed_date IS NOT NULL
    """)
    closed_rows = c.fetchall()

    # Compute business-day TAT for each closed issue
    tats = []
    eng_tats = {} # engineer_id -> list of their TATs
    for eng_id, a_str, c_str in closed_rows:
        a_dt = datetime.strptime(a_str, '%Y-%m-%d')
        c_dt = datetime.strptime(c_str, '%Y-%m-%d')
        bd = business_days_between(a_dt, c_dt, HOLIDAYS)
        tats.append(bd)
        eng_tats.setdefault(eng_id, []).append(bd)

    # 2. Team average
    team_avg = round(sum(tats) / len(tats), 2) if tats else 0.0

    # 3. Per-engineer stats: closed count, open count, avg TAT
    # First get open counts per engineer
    c.execute("""
        SELECT engineer_id,
               SUM(CASE WHEN status = 'Closed' THEN 1 ELSE 0 END) AS closed_cnt,
               SUM(CASE WHEN status <> 'Closed' THEN 1 ELSE 0 END) AS open_cnt
        FROM issues
        GROUP BY engineer_id
    """)
    counts = {row[0]: {'closed': row[1], 'open': row[2]} for row in c.fetchall()}

    # Now pull engineer names
    c.execute("SELECT engineer_id, name FROM engineers")
    names = {row[0]: row[1] for row in c.fetchall()}

    eng_stats = []
    for eng_id, name in names.items():
        closed_cnt = counts.get(eng_id, {}).get('closed', 0)
        open_cnt = counts.get(eng_id, {}).get('open', 0)
        tat_list = eng_tats.get(eng_id, [])
        avg_tat = round(sum(tat_list)/len(tat_list), 2) if tat_list else 0.0
        eng_stats.append((name, closed_cnt, open_cnt, avg_tat))

    conn.close()
    return team_avg, eng_stats



app = Flask(__name__)
app.secret_key = 'werysgvsretgseg'
CORS(app)

@app.route("/manager")
def manager_dashboard():
    name = request.args.get("name", "").strip()
  
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT manager_id FROM managers WHERE name = ?", (name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return f"No such manager '{name}'", 404
    mgr_id = row[0]
    c.execute("""
      SELECT jit_number, summary, status, created_date, assigned_date,
             (SELECT e.name FROM engineers e WHERE e.engineer_id = i.engineer_id) AS engineer,
             close_requested_by, closed_date, delete_issue,
             (SELECT t.name FROM technician t WHERE t.technician_id = i.technician_id) AS technician,
             severity, platform, close_requested_by
      FROM issues i
      WHERE ((status != 'Closed' AND created_date > DATETIME('now', '-1 day')) OR new_jit_add_date > DATETIME('now', '-1 day')) AND closed_date IS NULL
      ORDER BY created_date DESC
    """)
    new_issues = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, created_date, assigned_date,
             (SELECT e.name FROM engineers e WHERE e.engineer_id = i.engineer_id) AS engineer,
             close_requested_by, closed_date, delete_issue,
             (SELECT t.name FROM technician t WHERE t.technician_id = i.technician_id) AS technician,
             severity, platform, close_requested_by
      FROM issues i
      WHERE status != 'Closed' AND delete_issue == 0
      ORDER BY created_date DESC
      LIMIT 3
    """)
    issues = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, severity, assigned_date,
             (SELECT e.name FROM engineers e WHERE e.engineer_id = i.engineer_id) AS engineer,
             close_requested_by, closed_date, delete_issue,
             (SELECT t.name FROM technician t WHERE t.technician_id = i.technician_id) AS technician,
             platform
      FROM issues i
      WHERE status == 'Closed' AND new_jit_add_date IS NULL
      ORDER BY closed_date DESC
      LIMIT 3
    """)
    issues_closed = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, close_requested_by, close_requested_date, delete_issue,
              dam_used
      FROM issues i
      WHERE close_requested_by IS NOT NULL AND status != 'Closed' 
      ORDER BY closed_date DESC
    """)
    pending = c.fetchall()
    
    c.execute("SELECT COUNT(*) FROM issues WHERE status = 'Closed'")
    total_closed = c.fetchone()[0]
    c.execute("SELECT manager_id, name FROM managers")
    managers_list = c.fetchall()
    # new pending: list of (jit_number, category, value, introduced)
    c.execute("""
      SELECT jit_number, summary, status, close_requested_by, close_requested_date, delete_issue, dam_used
      FROM issues i
      WHERE close_requested_by IS NOT NULL AND status != 'Closed'
      ORDER BY close_requested_date DESC
    """)
    pending_issues = c.fetchall()
    # Fetch damage details
    c.execute("""
      SELECT i.jit_number, dt.category, dt.value, ed.introduced
      FROM issues i
      JOIN engineer_damage ed ON ed.issue_id = i.issue_id
      JOIN damage_types dt ON dt.id = ed.damage_id
      WHERE i.close_requested_by IS NOT NULL AND i.status != 'Closed'
      ORDER BY i.close_requested_date DESC
    """)
    damage_details = c.fetchall()
    from collections import defaultdict
    damage_by_issue = defaultdict(lambda: {'found': [], 'introduced': []})
    for jit, cat, val, introduced in damage_details:
        if introduced == 0:
            damage_by_issue[jit]['found'].append((cat, val))
        else:
            damage_by_issue[jit]['introduced'].append((cat, val))

    # Combine issue and damage data
    pending_closures = []
    for issue in pending_issues:
        jit = issue[0]
        damages = damage_by_issue.get(jit, {'found': [], 'introduced': []})
        pending_closures.append({
            'jit_number': jit,
            'summary': issue[1],
            'status': issue[2],
            'close_requested_by': issue[3],
            'close_requested_date': issue[4],
            'delete_issue': issue[5],
            'dam_used': issue[6],
            'found_damages': damages['found'],
            'introduced_damages': damages['introduced']
        })
    pending_closures_list = [{
        'jit_number': pc['jit_number'],
        'found': pc['found_damages'],
        'introduced': pc['introduced_damages'],
        'dam_used': pc['dam_used'],
        'summary': pc['summary']
    } for pc in pending_closures]

    # Serialize to JSON
    pending_closures_json = json.dumps(pending_closures_list)
    conn.close()
    team_avg, eng_stats = compute_metrics()
    return render_template(
        "manager_dashboard.html",
        manager=name,
        issues=issues,
        new_issues=new_issues,
       
        total_closed=total_closed,
        team_avg_tat=team_avg,
        eng_stats=eng_stats,
        managers_list=managers_list,
        issues_closed=issues_closed,
        pending=pending,
        pending_closures=pending_closures,
        pending_closures_json=pending_closures_json,
       
    )

@app.route("/assign_engineer")
def assign_engineer():
    issue_id = request.args.get("issue_id", "")
    manager = request.args.get("manager", "")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
      SELECT summary, status, created_date, severity
      FROM issues
      WHERE jit_number = ?
    """, (issue_id,))
    issue_row = c.fetchone()
    if not issue_row:
        conn.close()
        return f"Issue '{issue_id}' not found.", 404
    summary, status, created_date, severity = issue_row
    c.execute("""
      SELECT e.engineer_id, e.name,
             SUM(CASE WHEN i.status <> 'Closed' AND i.engineer_id = e.engineer_id THEN 1 ELSE 0 END) AS open_cnt,
             SUM(CASE WHEN i.status = 'Closed' AND i.engineer_id = e.engineer_id THEN 1 ELSE 0 END) AS closed_cnt
      FROM engineers e
      LEFT JOIN issues i ON i.engineer_id = e.engineer_id
      GROUP BY e.engineer_id
    """)
    engineers = c.fetchall()
    c.execute("""
      SELECT t.technician_id, t.name,
             SUM(CASE WHEN i.status <> 'Closed' AND i.technician_id = t.technician_id THEN 1 ELSE 0 END) AS open_cnt,
             SUM(CASE WHEN i.status = 'Closed' AND i.technician_id = t.technician_id THEN 1 ELSE 0 END) AS closed_cnt
      FROM technician t
      LEFT JOIN issues i ON i.technician_id = t.technician_id
      GROUP BY t.technician_id
    """)
    technician = c.fetchall()
    conn.close()
    return render_template(
        "assign_engineer.html",
        issue_id=issue_id,
        summary=summary,
        status=status,
        severity=severity,
        created_date=created_date,
        engineers=engineers,
        technician=technician,
        manager=manager
    )

@app.route("/edit_assign", methods=["POST"])
def edit_assign():
    issue_id = request.form.get("issue_id")
    manager_name = request.form.get("manager_name")
    if not issue_id or not manager_name:
        flash("Invalid form data.", "error")
        return redirect(f"/manager?name={manager_name or 'unknown'}")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
      SELECT engineer_id, close_requested_by
      FROM issues
      WHERE jit_number = ?
    """, (issue_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        flash(f"Issue {issue_id} not found.", "error")
        return redirect(f"/manager?name={manager_name}")
    old_engineer_id, close_requested_by = row
   
    if close_requested_by:
        conn.close()
        return render_template_string("""
        <script type="text/javascript">
        alert("Issue is requested to be closed! Cannot edit anymore.");
        window.location.href = "/manager?name={{manager_name}}";
        </script>
        """, manager_name=manager_name)
    if old_engineer_id:
        c.execute("SELECT name FROM engineers WHERE engineer_id = ?", (old_engineer_id,))
        engineer = c.fetchone()
       
        if engineer:
            message = f"Issue {issue_id} has been unassigned from you."
            c.execute("""
              INSERT INTO notifications (engineer_id, message, created_date,assigned_or_not)
              VALUES (?, ?, DATE('now'),'1')
            """, (old_engineer_id, message))
    c.execute("""
      UPDATE issues
      SET technician_id = NULL, engineer_id = NULL
      WHERE jit_number = ?
    """, (issue_id,))
    conn.commit()
    conn.close()
    return redirect(f"/assign_engineer?issue_id={issue_id}&manager={manager_name}")

@app.route("/do_assign", methods=["POST"])
def do_assign():
    issue_id = request.form["issue_id"]
    engineer_id = request.form.get("engineer_id")
    technician_id = request.form.get("technician_id")
    assignment_type = request.form.get("assignment_type")
    manager = request.form["manager"]
    if not issue_id or not manager or (assignment_type == "engineer" and not engineer_id) or (assignment_type == "technician" and not technician_id):
        flash("Invalid form data.", "error")
        return redirect(f"/manager?name={manager or 'unknown'}")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
      SELECT engineer_id, technician_id, engineer_history
      FROM issues
      WHERE jit_number = ?
    """, (issue_id,))
    result = c.fetchone()
    if not result:
        conn.close()
        flash(f"Issue {issue_id} not found.", "error")
        return redirect(f"/manager?name={manager}")
    old_engineer_id, old_technician_id, engineer_history = result
    if assignment_type == "engineer":
        # Update engineer_history
        history = json.loads(engineer_history or "[]")
        history.append(int(engineer_id))  # Ensure engineer_id is an integer
        new_history = json.dumps(history)
        c.execute("""
          UPDATE issues
          SET engineer_id = ?, assigned_date = DATE('now'), engineer_history = ?
          WHERE jit_number = ?
        """, (engineer_id, new_history, issue_id))
        
     
        message = f"You have been assigned to issue {issue_id}."
        c.execute("""
                  INSERT INTO notifications (engineer_id, message, created_date)
                  VALUES (?, ?, CURRENT_TIMESTAMP)
        """, (engineer_id, message))
    elif assignment_type == "technician":
        c.execute("""
          UPDATE issues
          SET technician_id = ?, assigned_date = DATE('now')
          WHERE jit_number = ?
        """, (technician_id, issue_id))
    conn.commit()
    c.execute("""
      SELECT engineer_id, technician_id
      FROM issues
      WHERE jit_number = ?
    """, (issue_id,))
    result = c.fetchone()
    conn.close()
    if result and result[0] and result[1]:
        return redirect(f"/manager?name={manager}")
    return redirect(f"/assign_engineer?issue_id={issue_id}&manager={manager}")

@app.route('/request_close', methods=["POST"])
def request_close():
    issue_id = request.form.get("issue_id")
    engineer = request.form.get("engineer")
    found_ids = request.form.getlist("found_ids[]")  # List of selected "found" damage IDs
    intro_ids = request.form.getlist("intro_ids[]")  # List of selected "introduced" damage IDs
    dam_used = 1 if request.form.get("dam_used") == '1' else 0
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # check existence & status as before
    c.execute("SELECT status, close_requested_by FROM issues WHERE jit_number = ?", (issue_id,))
    row = c.fetchone()
    if not row or row[0] == 'Closed' or row[1]:
        conn.close()
        return "Cannot request close"

    # update issue
    c.execute(
        "UPDATE issues SET close_requested_by=?, close_requested_date=?, dam_used=? WHERE jit_number=?",
        (engineer, now_str, dam_used, issue_id)
    )
    # clear old picks
    c.execute("DELETE FROM engineer_damage WHERE issue_id=(SELECT issue_id FROM issues WHERE jit_number=?)", (issue_id,))
    # insert new picks
    # found = introduced flag = 0
    for did in found_ids:
        c.execute(
            "INSERT INTO engineer_damage(issue_id, damage_id, introduced) VALUES ((SELECT issue_id FROM issues WHERE jit_number=?), ?, 0)",
            (issue_id, did)
        )
    # introduced = 1
    for did in intro_ids:
        c.execute(
            "INSERT INTO engineer_damage(issue_id, damage_id, introduced) VALUES ((SELECT issue_id FROM issues WHERE jit_number=?), ?, 1)",
            (issue_id, did)
        )
    conn.commit()
    conn.close()
    return redirect(url_for('engineer_dashboard', name=engineer))

@app.route("/confirm_close", methods=["POST"])
def confirm_close():
    issue_id = request.form.get("issue_id")
    manager_name = request.form.get("manager_name")
    comments = request.form.get("manager_comments")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT jit_number, status, close_requested_by
        FROM issues
        WHERE jit_number = ?
    """, (issue_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return "Issue not found", 404
    jit_number, status, close_req = row
    if status == "Closed" or not close_req:
        conn.close()
        return "Issue already closed or no close request pending", 404
    today_str = datetime.now().strftime("%Y-%m-%d")
    # Insert manager comment (assumes a manager_comments table exists)
    c.execute("INSERT INTO manager_comments (issue_id, comment) VALUES (?, ?)", (issue_id, comments))
    c.execute("""
        UPDATE issues
        SET status = 'Closed', closed_date = ?, new_jit_add_date = NULL
        WHERE jit_number = ?
    """, (today_str, issue_id))
    conn.commit()
    conn.close()
    update_excel_closure(jit_number, today_str)
    return redirect(url_for("manager_dashboard", name=manager_name))

def update_excel_closure(jit_number, today_str):
    import openpyxl
    workbook = openpyxl.load_workbook('new_jits.xlsx')
    worksheet = workbook.active
    search_id = jit_number
    column_to_update = 6
    new_value = 'Close'
    for row in worksheet.iter_rows():
        if row[0].value == search_id:
            worksheet.cell(row=row[0].row, column=column_to_update, value=new_value)
            break
    workbook.save('new_jits.xlsx')

@app.route("/reject_close",methods=["POST"])
def reject_close():

    issue_id=request.form.get("issue_id")
    manager_name=request.form.get("manager_name")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
  
    
    c.execute("""
        UPDATE issues SET close_requested_date =NULL , close_requested_by = NULL , reject_request =1,read1=0
        WHERE  jit_number=?
""",(issue_id,))
    conn.commit()
    c.execute("""SELECT reject_request FROM issues
              WHERE jit_number = ?""",(issue_id,))
    row = c.fetchone()
    print("issue_id is ",issue_id)
    print("reject_Request",row)

    conn.close()
    return redirect(url_for("manager_dashboard", name=manager_name))


@app.route("/delete_issue", methods=["POST"])
def delete_issue():
    issue_id = request.form.get("issue_id")
    manager_name = request.form.get("manager_name")
    if issue_id:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("UPDATE issues SET delete_issue = 1,engineer_id=NULL WHERE jit_number = ?", (issue_id,))
        conn.commit()
        conn.close()
    return redirect(url_for("manager_dashboard", name=manager_name))

@app.route("/all_issues")
def all_issues():
    manager = request.args.get("manager", "")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()


    c.execute("""
      SELECT jit_number, summary, status, created_date, assigned_date,
             (SELECT e.name FROM engineers e WHERE e.engineer_id = i.engineer_id) AS engineer,
             close_requested_by, closed_date, delete_issue,
             (SELECT t.name FROM technician t WHERE t.technician_id = i.technician_id) AS technician,
             severity, platform, core_cet
      FROM issues i
      WHERE (status != 'Closed' AND delete_issue == 0) OR   ( new_jit_add_date IS NOT NULL)
      ORDER BY created_date DESC
   
    """)
    issues = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, severity, assigned_date,
             (SELECT e.name FROM engineers e WHERE e.engineer_id = i.engineer_id) AS engineer,
             close_requested_by, closed_date, delete_issue,
             (SELECT t.name FROM technician t WHERE t.technician_id = i.technician_id) AS technician,
             platform,core_cet,created_date
      FROM issues i
      WHERE status = 'Closed' AND new_jit_add_date IS NULL
      ORDER BY closed_date DESC
  
    """)
    issues_closed = c.fetchall()

    c.execute("SELECT manager_id, name FROM managers")
    managers_list = c.fetchall()
    # new pending: list of (jit_number, category, value, introduced)
    c.execute("""
      SELECT jit_number, summary, status, close_requested_by, close_requested_date, delete_issue, dam_used
      FROM issues i
      WHERE close_requested_by IS NOT NULL AND status != 'Closed'
      ORDER BY close_requested_date DESC
    """)
    pending_issues = c.fetchall()
    # Fetch damage details
    c.execute("""
      SELECT i.jit_number, dt.category, dt.value, ed.introduced
      FROM issues i
      JOIN engineer_damage ed ON ed.issue_id = i.issue_id
      JOIN damage_types dt ON dt.id = ed.damage_id
      WHERE i.close_requested_by IS NOT NULL AND i.status != 'Closed'
      ORDER BY i.close_requested_date DESC
    """)
    damage_details = c.fetchall()
    from collections import defaultdict
    damage_by_issue = defaultdict(lambda: {'found': [], 'introduced': []})
    for jit, cat, val, introduced in damage_details:
        if introduced == 0:
            damage_by_issue[jit]['found'].append((cat, val))
        else:
            damage_by_issue[jit]['introduced'].append((cat, val))

    # Combine issue and damage data
    pending_closures = []
    for issue in pending_issues:
        jit = issue[0]
        damages = damage_by_issue.get(jit, {'found': [], 'introduced': []})
        pending_closures.append({
            'jit_number': jit,
            'summary': issue[1],
            'status': issue[2],
            'close_requested_by': issue[3],
            'close_requested_date': issue[4],
            'delete_issue': issue[5],
            'dam_used': issue[6],
            'found_damages': damages['found'],
            'introduced_damages': damages['introduced']
        })
    pending_closures_list = [{
        'jit_number': pc['jit_number'],
        'found': pc['found_damages'],
        'introduced': pc['introduced_damages'],
        'dam_used': pc['dam_used'],
        'summary': pc['summary']
    } for pc in pending_closures]

    # Serialize to JSON
    pending_closures_json = json.dumps(pending_closures_list)

    conn.close()
    return render_template(
        "all_issues.html",
        manager=manager,
        issues=issues,
         issues_closed=issues_closed,
        managers_list=managers_list,
       
        pending_closures=pending_closures,
        pending_closures_json=pending_closures_json,
    )

@app.route("/mark_notifications_as_read", methods=["POST"])
def mark_notifications_as_read():
    engineer_name = request.form.get("engineer_name")
    print("engineer name is ",engineer_name)
    if not engineer_name:
        return jsonify({"success": False, "message": "Engineer name is required."}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("SELECT engineer_id FROM engineers WHERE name = ?", (engineer_name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "Engineer not found."}), 404
    eng_id = row[0]
    c.execute("UPDATE notifications SET read1 = 1 WHERE engineer_id = ? AND read1 = 0", (eng_id,))
    c.execute("UPDATE issues SET read1 = 1 WHERE engineer_id = ? AND read1 = 0", (eng_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/mark_notifications_as_read2", methods=["POST"])
def mark_notifications_as_read2():
    technician_name = request.form.get("technician_name")
    print("engineer name is ",technician_name)
    if not technician_name:
        return jsonify({"success": False, "message": "Engineer name is required."}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("SELECT technician_id FROM technician WHERE name = ?", (technician_name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "Engineer not found."}), 404
    tech_id = row[0]
    c.execute("UPDATE notifications SET read2 = 1 WHERE technician_id = ? AND read2 = 0", (tech_id,))
    c.execute("UPDATE issues SET read2 = 1 WHERE technician_id = ? AND read2 = 0", (tech_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})
    


def fetch_jit_details(jit_number):
    try:
        issue = jira.issue(jit_number)
        fields = issue.fields
        platform = fields.customfield_20012.value if fields.customfield_20012 else "-"
        core = platform_to_corecet.get(platform.strip().upper(), 'Unknown')
        component = fields.customfield_18711.value if fields.customfield_18711 else "-"
        root_cause = fields.customfield_15203 if fields.customfield_15203 else "-"
        status = fields.status.name if fields.status else "-"
        created_date = fields.created.split("T")[0] if fields.created else "-"
        severity = fields.customfield_10005.value.split('-')[0].strip() if fields.customfield_10005 and fields.customfield_10005.value and '-' in fields.customfield_10005.value else fields.customfield_10005.value if fields.customfield_10005 and fields.customfield_10005.value else "-"
        comments = issue.fields.comment.comments if issue.fields.comment else []
        rce_assigned_engineer = ", ".join([name for name in tracked_names if any(name in comment.body for comment in comments)]) or "-"
        return {
            "JIT Number": issue.key,
            "CORE-CET": core,
            "Platform": platform,
            "Summary": fields.summary,
            "Status": "Open",
            "RCE Assigned Engineer": rce_assigned_engineer,
            "technician": None,
            "Severity": severity,
            "Created Date": created_date,
            "Component": component,
            "Root Cause": root_cause,
        }
    except Exception as e:
        print(f"Error fetching JIT details: {e}")
        return None

@app.route('/add_jit', methods=["POST"])
def add_jit():
    jit_number = request.form.get("jit_number", '').strip()
    manager_name = request.args.get('name') or request.form.get('manager_name')
    details = fetch_jit_details(jit_number)
    if not details:
        flash(f"JIT {jit_number} not found in Jira", "error")
        return redirect(url_for('manager_dashboard', name=manager_name))
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT manager_id FROM managers WHERE name = ?", (manager_name,))
    mgr = c.fetchone()[0]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        INSERT INTO issues (jit_number, summary, status, created_date, severity, platform, manager_id, new_jit_add_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(jit_number) DO NOTHING
    """, (
        details['JIT Number'],
        details['Summary'],
        details['Status'],
        details['Created Date'],
        details['Severity'],
        details['Platform'],
        mgr,
        now_str
    ))
    conn.commit()
    conn.close()
    flash(f"jit number {jit_number} has been added successfully")
    return redirect(url_for('manager_dashboard', name=manager_name))



def mark_as_read():
    name =request.args.get("name")

@app.route("/engineer")
def engineer_dashboard():
    name = request.args.get("name", "").strip()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT engineer_id FROM engineers WHERE name = ?", (name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return f"No such engineer '{name}'", 404
    eng_id = row[0]

    c.execute("""
      SELECT jit_number, summary, status, assigned_date,closed_date,close_requested_date
            
      FROM issues 
      WHERE status != 'Closed' AND assigned_date > DATETIME('now', '-1 day') AND engineer_id=?
      ORDER BY created_date DESC
    """,(eng_id,))
    new_issues = c.fetchall()

    c.execute("""
      SELECT jit_number, summary, status, assigned_date, closed_date, close_requested_date
      FROM issues
      WHERE engineer_id = ? AND assigned_date < DATETIME('now','-1 day')
      ORDER BY created_date DESC
    """, (eng_id,))
    issues = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, assigned_date, closed_date, close_requested_date
      FROM issues
      WHERE engineer_id = ? AND status == "Closed"
      ORDER BY created_date DESC
    """, (eng_id,))
    closed_issues = c.fetchall()
    c.execute("""
      SELECT message, created_date, read1
      FROM notifications
      WHERE engineer_id = ? AND read1 = 0 AND assigned_or_not= 1
      ORDER BY created_date DESC
    """, (eng_id,))
    notifications = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, assigned_date
      FROM issues i
      WHERE engineer_id = ? AND status != 'Closed' AND close_requested_date IS NULL AND read1==0
      ORDER BY assigned_date DESC
    """, (eng_id,))
    pending = c.fetchall()
    team_avg, eng_stats = compute_metrics()
    own_stats = next((stats for stats in eng_stats if stats[0] == name), None)

    #notification for rejected request
    c.execute("""
      SELECT jit_number
      FROM issues
      WHERE engineer_id = ?  AND reject_request=1 AND read1==0
      ORDER BY created_date DESC
    """, (eng_id,))
    rejected_request=c.fetchall()
   
    # Fetch all damage options
    c.execute("""
    SELECT MIN(id) AS id, category, value
    FROM damage_types
    GROUP BY category, value
    ORDER BY category, value
""")
    damages = c.fetchall()  # list of tuples (id, category, value)
    damages_by_cat = {}
    for d_id, cat, val in damages:
        damages_by_cat.setdefault(cat.strip(), []).append((d_id, val.strip()))
        print("is it repeating", d_id, cat, val)
    conn.close()
    return render_template(
        "engineer_dashboard.html",
        engineer=name,
        issues=issues,
        team_avg_tat=team_avg,
        own_stats=own_stats,
        pending_requests=pending,
        notifications=notifications,
        rejected_request=rejected_request,
        new_issues=new_issues,
        closed_issues=closed_issues,
        damages=damages,
        damage_by_category=damages_by_cat,
    )

@app.route('/technician')
def technician():
    name = request.args.get("name", "").strip()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT technician_id FROM technician WHERE name = ?", (name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return f"No such technician '{name}'", 404
    tech_id = row[0]
    c.execute("""
      SELECT jit_number, summary, status, assigned_date,closed_date,close_requested_date
            
      FROM issues 
      WHERE status != 'Closed' AND assigned_date > DATETIME('now', '-1 day') AND engineer_id=?
      ORDER BY created_date DESC
    """,(tech_id,))
    new_issues = c.fetchall()

    c.execute("""
      SELECT jit_number, summary, status, created_date, closed_date
      FROM issues
      WHERE technician_id = ? AND assigned_date < DATETIME('now', '-1 day') 
      ORDER BY created_date DESC
    """, (tech_id,))
    issues = c.fetchall()
    c.execute("""
      SELECT jit_number, summary, status, assigned_date, closed_date, close_requested_date
      FROM issues
      WHERE engineer_id = ? AND status == "Closed"
      ORDER BY created_date DESC
    """, (tech_id,))
    closed_issues = c.fetchall()

    c.execute("""
      SELECT jit_number, summary, status, assigned_date
      FROM issues i
      WHERE technician_id = ? AND status != 'Closed' AND close_requested_date IS NULL AND read2=0
      ORDER BY assigned_date DESC
    """, (tech_id,))
    pending = c.fetchall()
    conn.close()
    return render_template(
        "technician_dashboard.html",
        technician=name,
        issues=issues,
        pending_requests=pending,
        new_issues=new_issues,
        closed_issues=closed_issues,
    )

@app.route("/team")
def team():
    manager = "Mohanraj_Chinnasamy" or "Pravin_Janakiram"
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 1. Fetch all CLOSED issues with their assigned and closed dates
    c.execute("""
        SELECT engineer_id, assigned_date, closed_date
        FROM issues
        WHERE status = 'Closed'
          AND assigned_date IS NOT NULL
          AND closed_date IS NOT NULL
    """)
    closed_rows = c.fetchall()
    
     # Compute business-day TAT for each closed issue
    tats = []
    eng_tats = {} # engineer_id -> list of their TATs
    for eng_id, a_str, c_str in closed_rows:
        a_dt = datetime.strptime(a_str, '%Y-%m-%d')
        c_dt = datetime.strptime(c_str, '%Y-%m-%d')
        bd = business_days_between(a_dt, c_dt, HOLIDAYS)
        tats.append(bd)
        eng_tats.setdefault(eng_id, []).append(bd)
         # First get open counts per engineer
    c.execute("""
        SELECT engineer_id,
               SUM(CASE WHEN status = 'Closed' THEN 1 ELSE 0 END) AS closed_cnt,
               SUM(CASE WHEN status <> 'Closed' THEN 1 ELSE 0 END) AS open_cnt
        FROM issues
        GROUP BY engineer_id
    """)
    counts = {row[0]: {'closed': row[1], 'open': row[2]} for row in c.fetchall()}

    # Now pull engineer names
    c.execute("SELECT engineer_id, name FROM engineers")
    names = {row[0]: row[1] for row in c.fetchall()}

    eng_stats = []
    for eng_id, name in names.items():
        closed_cnt = counts.get(eng_id, {}).get('closed', 0)
        open_cnt = counts.get(eng_id, {}).get('open', 0)
        tat_list = eng_tats.get(eng_id, [])
        avg_tat = round(sum(tat_list)/len(tat_list), 2) if tat_list else 0.0
        eng_stats.append((name,  open_cnt,closed_cnt, avg_tat))
   
    conn.close()
    return render_template("team.html", engineers=eng_stats, manager=manager)

# stats new function

# Endpoint for core/cet dropdown options
@app.route('/core_cet_options', methods=['GET'])
def get_core_cet_options():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT core_cet FROM issues")
    options = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return jsonify(options)

# Endpoint for platform dropdown options
@app.route('/platforms', methods=['GET'])
def get_platforms():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT platform FROM issues")
    options = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return jsonify(options)

# Endpoint for generation dropdown options
"""@app.route('/generations', methods=['GET'])
def get_generations():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT generation FROM issues")
    options = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return jsonify(options)"""

# Endpoint for issue counts based on selected combination
@app.route('/issue_counts', methods=['GET'])
def get_issue_counts():
    # Get query parameters from the frontend
    core_cet = request.args.get('core_cet')
    platform = request.args.get('platform')
   # generation = request.args.get('generation')

    # Validate that all parameters are provided
    if not all([core_cet, platform]):
        return jsonify({'error': 'Missing parameters'}), 400

    # Connect to the database and execute query
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = """
    SELECT
        SUM(CASE WHEN status = 'Closed' THEN 1 ELSE 0 END) AS closed,
        SUM(CASE WHEN status = 'Open' THEN 1 ELSE 0 END) AS open
    FROM issues
    WHERE core_cet = ? AND platform = ? 
    """
    cursor.execute(query, (core_cet, platform))
    result = cursor.fetchone()

    # Clean up database resources
    cursor.close()
    conn.close()

    # Process query result
    if result:
        closed, open_ = result
        return jsonify({'closed': closed or 0, 'open': open_ or 0})
    return jsonify({'closed': 0, 'open': 0})


if __name__ == "__main__":
    init_db()
    new_jits_df = pd.read_excel("new_jits.xlsx")
    new_jits_df["Status"] = "Open"
    new_jits_df["RCE Assigned Engineer"] = None
    process_new_jits(new_jits_df)
    app.run(debug=True)